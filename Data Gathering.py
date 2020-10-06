# -*- coding: utf-8 -*-
"""
Created on Sat Dec 21 21:17:03 2019

@author: berke
"""
from pykinect2 import PyKinectV2
from pykinect2 import PyKinectRuntime

import pygame
import ctypes
import sys

import numpy as np
import os
import pandas as pd
import math
from openpyxl import load_workbook

if sys.hexversion >= 0x03000000:
    import _thread as thread
else:
    import thread


data_folder = "data2"
excel_path = "./data2/body_joints2.xlsx"
image_folder = "./data2/images"
exclude_header_indexes = [3, 15, 19, 21, 22, 23, 24]
exclude_attribute_indexes = [5, 6, 7, 8]
 
SKELETON_COLORS = [pygame.color.THECOLORS["red"],
                  pygame.color.THECOLORS["blue"], 
                  pygame.color.THECOLORS["green"], 
                  pygame.color.THECOLORS["orange"], 
                  pygame.color.THECOLORS["purple"], 
                  pygame.color.THECOLORS["yellow"], 
                  pygame.color.THECOLORS["violet"]]


class KinectRuntime(object):
    def __init__(self):
        self._kinect = PyKinectRuntime.PyKinectRuntime(PyKinectV2.FrameSourceTypes_Color
                                                       | PyKinectV2.FrameSourceTypes_Body 
                                                       | PyKinectV2.FrameSourceTypes_Depth)
        
        self._bodies = None
        self._done = False
        self._filename = os.listdir(image_folder)
        self._counter = len(self._filename)
        self._active = False
        self._threshold = 10
        self._width_color_frame = self._kinect.color_frame_desc.Width
        self._height_color_frame = self._kinect.color_frame_desc.Height
        self._width_depth_frame = self._kinect.depth_frame_desc.Width 
        self._height_depth_frame = self._kinect.depth_frame_desc.Height
        
        pygame.init()
        pygame.display.set_caption("Gathering skeletal data")
        self._clock = pygame.time.Clock()
        self._infoObject = pygame.display.Info()
        self._frame_surface = pygame.Surface((self._width_color_frame, self._height_color_frame), 0, 32)
        self._screen = pygame.display.set_mode((self._infoObject.current_w >> 1, self._infoObject.current_h >> 1),
                                               pygame.HWSURFACE|pygame.DOUBLEBUF|pygame.RESIZABLE, 32)
        
        CSP_array = PyKinectV2._CameraSpacePoint * (self._width_depth_frame * self._height_depth_frame)
        self._depth_to_CSP_buffer = CSP_array()

        
    def appendDF2excel(self, filename, DF, count, sheet_name='Sheet', startrow=None, truncate_sheet=False,
                           **to_excel_kwargs):

        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')
            
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        writer.book = load_workbook(filename)
        
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        if count == 0:
            startrow = 1
            
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}    
        DF.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
        writer.save()
    
    def mapDepthFrame2CSP(self, depth_frame):
        length = self._width_depth_frame * self._height_depth_frame
        assert length == depth_frame.size
        ptr_depth = np.ctypeslib.as_ctypes(depth_frame.flatten())
        error_state = self._kinect._mapper.MapDepthFrameToCameraSpace(
                                                                length, ptr_depth,
                                                                length, self._depth_to_CSP_buffer)

        if error_state:
            raise "Could not map depth frame to camera space! " + str(error_state)
            
        pf_CSPs = ctypes.cast(self._depth_to_CSP_buffer, ctypes.POINTER(ctypes.c_float))
        data = np.ctypeslib.as_array(pf_CSPs, shape=(self._height_depth_frame, self._width_depth_frame, 3))
        
        del pf_CSPs, ptr_depth, depth_frame
        return np.copy(data)
    
    def draw_body_bone(self, joints, jointPoints, color, joint0, joint1):
        joint0State = joints[joint0].TrackingState;
        joint1State = joints[joint1].TrackingState;

        if (joint0State == PyKinectV2.TrackingState_NotTracked) or (joint1State == PyKinectV2.TrackingState_NotTracked): 
            return
        
        start = (jointPoints[joint0].x, jointPoints[joint0].y)
        end = (jointPoints[joint1].x, jointPoints[joint1].y)
        try:
            pygame.draw.line(self._frame_surface, color, start, end, 8)  
        except:
            pass
        
    def drawBody(self, joints, jointPoints, color):
        # Torso
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_Head, PyKinectV2.JointType_Neck);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_Neck, PyKinectV2.JointType_SpineShoulder);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_SpineShoulder, PyKinectV2.JointType_SpineMid);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_SpineMid, PyKinectV2.JointType_SpineBase);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_SpineShoulder, PyKinectV2.JointType_ShoulderRight);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_SpineShoulder, PyKinectV2.JointType_ShoulderLeft);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_SpineBase, PyKinectV2.JointType_HipRight);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_SpineBase, PyKinectV2.JointType_HipLeft);
    
        # Right Arm    
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_ShoulderRight, PyKinectV2.JointType_ElbowRight);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_ElbowRight, PyKinectV2.JointType_WristRight);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_WristRight, PyKinectV2.JointType_HandRight);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_HandRight, PyKinectV2.JointType_HandTipRight);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_WristRight, PyKinectV2.JointType_ThumbRight);

        # Left Arm
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_ShoulderLeft, PyKinectV2.JointType_ElbowLeft);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_ElbowLeft, PyKinectV2.JointType_WristLeft);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_WristLeft, PyKinectV2.JointType_HandLeft);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_HandLeft, PyKinectV2.JointType_HandTipLeft);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_WristLeft, PyKinectV2.JointType_ThumbLeft);

        # Right Leg
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_HipRight, PyKinectV2.JointType_KneeRight);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_KneeRight, PyKinectV2.JointType_AnkleRight);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_AnkleRight, PyKinectV2.JointType_FootRight);

        # Left Leg
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_HipLeft, PyKinectV2.JointType_KneeLeft);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_KneeLeft, PyKinectV2.JointType_AnkleLeft);
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_AnkleLeft, PyKinectV2.JointType_FootLeft);
        
    def drawColorFrame(self, RGB_frame, target_surface):
        target_surface.lock()
        address = self._kinect.surface_as_array(target_surface.get_buffer())
        ctypes.memmove(address, RGB_frame.ctypes.data, RGB_frame.size)
        del address
        target_surface.unlock()
    
    def getDepthFrameInfo(self, depth_frame, depth_joint_points):
        depth_joints_data = []
        for joint in list(depth_joint_points):
            X = round(joint.x)
            Y = round(joint.y)
            Z = depth_frame[Y * self._width_depth_frame + X]
            depth_joints_data.append((X, Y, Z))
        return depth_joints_data
    
    def getRGBframeInfo(self, RGB_joint_points):
        RGB_joints_data = []
        for joint in list(RGB_joint_points):
            X = round(joint.x)
            Y = round(joint.y)
                
            RGB_joints_data.append((X, Y))
        return RGB_joints_data
    
    def getRealJointCoordinates(self, depth_frame, depth_joint_points):
        CSP = self.mapDepthFrame2CSP(depth_frame) 
        real_joints_data = []
        for joint in list(depth_joint_points):
            X = CSP[round(joint.y)][round(joint.x)][0]
            Y = CSP[round(joint.y)][round(joint.x)][1]
            Z = CSP[round(joint.y)][round(joint.x)][2]
            real_joints_data.append((X, Y, Z))
        return real_joints_data
    
    def getJointQuaternions(self, orientations):
        
        quats = []        
        for joint in range(PyKinectV2.JointType_Count):            
            Qw = orientations[joint].Orientation.w
            Qx = orientations[joint].Orientation.x
            Qy = orientations[joint].Orientation.y
            Qz = orientations[joint].Orientation.z
            quats.append((Qw, Qx, Qy, Qz))
        return quats
    
    def getQuatsVectors(self, quaternions):
        quatsVectors = []
        for joint in range(PyKinectV2.JointType_Count):
            Quat = quaternions[joint] 
            
            Qw = Quat[0] 
            Qx = Quat[1]
            Qy = Quat[2]
            Qz = Quat[3]
            
            W = math.degrees(math.acos(Qw)) * 2
            X = Qx / math.sin(math.radians(W/2))
            Y = Qy / math.sin(math.radians(W/2))
            Z = Qz / math.sin(math.radians(W/2))
            
            quatsVectors.append((W, X, Y, Z))
        return quatsVectors
    
    def getDFinfo(self, RGB_joints_data, real_joints_data, orientation_vectors):
        rowInfo = []
        
        for joint in range(PyKinectV2.JointType_Count):
            rowInfo.append(RGB_joints_data[joint][0])
            rowInfo.append(RGB_joints_data[joint][1])
            rowInfo.append(real_joints_data[joint][0])
            rowInfo.append(real_joints_data[joint][1])
            rowInfo.append(real_joints_data[joint][2])
            
            if joint not in exclude_header_indexes:
                rowInfo.append(orientation_vectors[joint][0])
                rowInfo.append(orientation_vectors[joint][1])
                rowInfo.append(orientation_vectors[joint][2])
                rowInfo.append(orientation_vectors[joint][3])
        
        rowInfo.append(23) # recorded persons id
        excel_row = []
        excel_row.append(rowInfo)
        DF = pd.DataFrame(excel_row)
        return DF
    
    
    def checkActiveState(self, RGB_joint_points):
        if RGB_joint_points[11].x or RGB_joint_points[11].y or RGB_joint_points[7].x or RGB_joint_points[7].y != float("inf"):          
            r_hand_X = round(RGB_joint_points[11].x)
            r_hand_Y = round(RGB_joint_points[11].y)
            l_hand_X = round(RGB_joint_points[7].x)
            l_hand_Y = round(RGB_joint_points[7].y)
        else:
            self._active = False
            return self._active
                                
        if (r_hand_X - self._threshold < l_hand_X < r_hand_X + self._threshold) and \
            (r_hand_Y - self._threshold < l_hand_Y < r_hand_X + self._threshold):
                if self._active:
                    self._active = False
                else:
                    self._active = True
        return self._active
            
      
    def Runtime(self):
        while not self._done:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    self._done = True

                elif event.type == pygame.VIDEORESIZE:
                    self._screen = pygame.display.set_mode(event.dict['size'], 
                                               pygame.HWSURFACE|pygame.DOUBLEBUF|pygame.RESIZABLE, 32)

            if self._kinect.has_new_color_frame():
                RGB_frame = self._kinect.get_last_color_frame()
                depth_frame = self._kinect.get_last_depth_frame()
                self._bodies = self._kinect.get_last_body_frame()
                
                self.drawColorFrame(RGB_frame, self._frame_surface)
                if self._bodies is not None: 
                    for i in range(0, self._kinect.max_body_count):
                        body = self._bodies.bodies[i]
                        if body.is_tracked:
                            joints = body.joints
                            RGB_joint_points = self._kinect.body_joints_to_color_space(joints)
                            #active_state = self.checkActiveState(RGB_joint_points)
                            
                            orientations = body.joint_orientations
                            
                            depth_joint_points = self._kinect.body_joints_to_depth_space(joints)
                            quaternions = self.getJointQuaternions(orientations)

                            RGB_joints_data = self.getRGBframeInfo(RGB_joint_points)
                            #depth_joints_data = self.getDepthFrameInfo(depth_frame, depth_joint_points)
                            real_joints_data = self.getRealJointCoordinates(depth_frame, depth_joint_points)
                            orientation_vectors = self.getQuatsVectors(quaternions)
                            
                            DF = self.getDFinfo(RGB_joints_data, real_joints_data, orientation_vectors)
                            self.appendDF2excel(excel_path, DF, count=self._counter, header=None, index=False, sheet_name='Sheet')                        
                            
                            self.drawBody(joints, RGB_joint_points, SKELETON_COLORS[i])
                            self._counter += 1
                            pygame.image.save(self._frame_surface, "data2\images\RGBskeleton_"+ str(self._counter)+".jpeg")                                        

            h_to_w = float(self._frame_surface.get_height()) / self._frame_surface.get_width()
            target_height = int(h_to_w * self._screen.get_width())
            surface_to_draw = pygame.transform.scale(self._frame_surface, (self._screen.get_width(), target_height));
            self._screen.blit(surface_to_draw, (0,0))
            surface_to_draw = None
            
            pygame.display.update()
            pygame.display.flip()
            self._clock.tick(15) #FPS

        self._kinect.close()
        pygame.quit()

def setHeader(headers):
    attributes = ['_IMG_X', '_IMG_Y',
                  '_koord_X', '_koord_Y', '_koord_Z',
                  '_aci_W', '_vek_X', '_vek_Y', '_vek_Z']
    excel_header = []

    empty_header_data = [[]]
    empty_header_data[0].append(0)
    
    for i in range(len(headers)):
        for j in range(len(attributes)):
            if not( i in exclude_header_indexes and j in exclude_attribute_indexes ):
                empty_header_data[0].append(0)
                atr = headers[i] + attributes[j]
                excel_header.append(atr)
    excel_header.append("Kişi")
    
    empty_DF = pd.DataFrame(empty_header_data)
    empty_DF.to_excel(excel_path, sheet_name='Sheet', index=None, header=excel_header, startrow=0)
    return empty_DF
    
if __name__ == "__main__":
    if not os.path.exists(data_folder):
        os.makedirs(data_folder)
    if not os.path.exists(image_folder):
        os.makedirs(image_folder)
    if not os.path.exists(excel_path):
        joint_headers = ['Omurga_Tabanı', 'Orta_Omurga', 'Boyun', 'Bas',
                        'Sol_Omuz', 'Sol_Dirsek', 'Sol_Bilek', 'Sol_El',
                        'Sag_Omuz', 'Sag_Dirsek', 'Sag_Bilek', 'Sag_El',
                        'Sol_Kalca', 'Sol_Diz', 'Sol_Ayak_Bilegi', 'Sol_Ayak',
                        'Sag_Kalca', 'Sag_Diz', 'Sag_Ayak_Bilegi', 'Sag_Ayak',
                        'Ust_Omurga',
                        'Sol_El_Ucu', 'Sol_Bas_Parmak',
                        'Sag_El_Ucu', 'Sag_Bas_Parmak']
        excel_DF = setHeader(joint_headers)
        wb = load_workbook(excel_path)
        sheet = wb.active
        cols = sheet.columns
        for col in cols:
            sheet.column_dimensions[col[0].column_letter].width = 23
        wb.save(excel_path)
        
    
    test = KinectRuntime();
    test.Runtime();

